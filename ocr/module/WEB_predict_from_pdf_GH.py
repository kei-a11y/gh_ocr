import os, io, sys,re, cv2#, pytesseract
from pathlib import Path
from PIL import Image, ImageOps, ImageEnhance
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from django.conf import settings


#pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

# --- 外部ライブラリ読み込み（無ければ適宜メッセージ） ---
try:
    import fitz  # PyMuPDF
except Exception:
    raise RuntimeError("PyMuPDF が必要です。pip install pymupdf を実行してください。")

try:
    import cv2
    _have_cv2 = True
except Exception:
    _have_cv2 = False
    print("WARN: OpenCV が見つかりません。桁分割は簡易モードになります。pip install opencv-python を推奨します.")

try:
    import tensorflow as tf
    _have_tf = True
except Exception:
    _have_tf = False
    print("WARN: tensorflow が見つかりません。モデル推論は実行されません。pip install tensorflow を実行してください.")

# ------------------ ユーザー設定エリア ------------------
BASE_DIR = Path(__file__).parent # スクリプトのあるディレクトリを取得
PDF_PATH = BASE_DIR / "@pdf" / "kk.pdf"
PAGE_INDEX = 0
DEBUG_DIR = "debug_crops"
DPI = 200

# 正規化座標 (ユーザーが調整済みの値をそのまま使用)　★左端起点x0,上端起点y0,左端終点x1,上端終点y1
BOX_RECIPIENT = (0.145, 0.100, 0.285, 0.135)
BOX_NAME = (0.435, 0.100, 0.595, 0.1175)
BOX_RESIDENCE = (0.410, 0.145, 0.435, 0.165)
BOX_SERVICE   = (0.1435, 0.2225, 0.1675, 0.8455)
BOX_JUTOGAI   = (0.2425, 0.2225, 0.2755, 0.8455)
BOX_NIGHT     = (0.3135, 0.2225, 0.3595, 0.8455)
BOX_HOSPITA   = (0.3615, 0.2225, 0.4085, 0.8455)
BOX_GOhome    = (0.4095, 0.2225, 0.4575, 0.8455)
BOX_SUPPORT   = (0.4585, 0.2225, 0.5065, 0.8455)
BOX_MEDIC     = (0.5075, 0.2225, 0.5545, 0.8455)
BOX_JIRITSU1  = (0.5565, 0.2225, 0.6025, 0.8455)
BOX_JIRITSU2  = (0.6045, 0.2225, 0.6515, 0.8455)

NUM_ROWS = 31
MODEL_CANDIDATES = ["best_model.h5", "mnist_model.h5"]
# -------------------------------------------------------

def ensure_dir(p):
    Path(p).mkdir(parents=True, exist_ok=True)

#--- PDFの指定ページを画像（PIL形式）に変換する関数
def pdf_page_to_pil(pdf_path, page_num=0, dpi=200):
    doc = fitz.open(pdf_path)
    if page_num < 0 or page_num >= doc.page_count:
        raise ValueError("page_num out of range")
    page = doc.load_page(page_num)
    mat = fitz.Matrix(dpi / 72, dpi / 72)
    pix = page.get_pixmap(matrix=mat) #-----------------PDFページをピクセルマップに変換
    img = Image.open(io.BytesIO(pix.tobytes("png"))) #--PIL画像に変換
    return img.convert("RGB")


#--- 画像を指定された範囲で切り取るための関数
def crop_norm(img, box):#--- img=PILのImageオブジェクト、box=正規化された座標（main内で代入）
    w, h = img.size
    x0 = int(box[0] * w)
    y0 = int(box[1] * h)
    x1 = int(box[2] * w)
    y1 = int(box[3] * h)
    return img.crop((x0, y0, x1, y1)) #--- PILのcrop()メソッドを使って指定範囲を切り取る


#---31日だから31分割するコード（罫線を検出できなかったときのためのフォールバック）
def split_vertical(box_img, n_rows):
    w, h = box_img.size
    row_h = h / n_rows
    crops = []
    for i in range(n_rows):
        y0 = int(i * row_h)
        y1 = int((i + 1) * row_h)
        crops.append(box_img.crop((0, y0, w, y1)))
    return crops


#---罫線を検出して区切るコード　#---エクセルの罫線は太くしないと認識しない！
def split_rows_by_lines(img, debug_name=""):
    gray = np.array(img.convert("L"))
    _, bw = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)

    # 横線を強調（横方向に長い形状を抽出）
    kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (50, 1))
    horizontal = cv2.morphologyEx(bw, cv2.MORPH_OPEN, kernel, iterations=2)

    # 輪郭から罫線候補を検出
    contours, _ = cv2.findContours(horizontal, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    y_positions = []
    for c in contours:
        x, y, w, h = cv2.boundingRect(c)
        if w > img.width * 0.5:  # ページ幅の半分以上なら罫線とみなす
            y_positions.append(y)

    if not y_positions:
        print("Warning: 罫線が検出できませんでした")
        return [img]

    # y座標をソート
    y_positions = sorted(y_positions)

    # 隣り合う罫線の間を切り出し
    rows = []
    for i in range(len(y_positions) - 1):
        y0, y1 = y_positions[i], y_positions[i+1]
        row_img = img.crop((0, y0, img.width, y1))
        if debug_name:
            ensure_dir("mid")
            row_img.save(Path("mid") / f"{debug_name}_row_{i+1}.png")
        rows.append(row_img)

    return rows



#---受給者証番号をOCR + fallbackで認識
def predict_recipient_number(img, model, debug_name="recipient"):
    """受給者証番号 → モデルのみで予測"""
    if model is not None:
        val = predict_digits_with_model(model, img, debug_name)
        return val or ""
    return ""


#---受給者名（カナ）をOCRで認識
def predict_name(img, model=None, debug_name="name"):
    """受給者名（カナ） → ここでは未対応（空文字に）"""
    # Tesseractを使わないので、常に空白にしておく
    return ""


#---住居番号をOCR + fallbackで認識
def predict_residence_number(img, model, debug_name="residence"):
#    import re, pytesseract

#    # --- Step 1: Tesseract OCR (数字限定) ---
#    config = "--psm 7 -c tessedit_char_whitelist=0123456789"
#    text = pytesseract.image_to_string(img, config=config).strip()
#    digits = re.sub(r"\D", "", text)  # 数字以外削除

#    # --- Step 2: Tesseractの結果が10桁なら即採用 ---
#    if len(digits) == 1:
#        return digits

    # --- Step 3: fallback: 手書きモデル ---
    if model is not None:
        val = predict_digits_with_model(model, img, debug_name)
        # → 桁数はとりあえず返す（Excelに痕跡を残す）
        if val:
            return val or ""
        return ""
#    # --- Step 4: どちらも失敗 ---
#    return digits  # 



#--- 四隅のランドマーク(▣)を基準に射影変換
def align_using_corner_marks(scanned_img, dpi, template_pdf_path):
    import cv2
    import numpy as np
    from pathlib import Path
    from PIL import Image
    
    # テンプレート画像化
    template_img_path = "template_image.png"
    if not Path(template_img_path).exists():
        tmp_img = pdf_page_to_pil(template_pdf_path, 0, dpi)
        tmp_img.save(template_img_path)
    template_img = cv2.cvtColor(np.array(Image.open(template_img_path)), cv2.COLOR_RGB2BGR)
    
    def find_marks(img_bgr, debug=False):
        gray = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2GRAY)
        
        # より強力な前処理
        # 1. ガウシアンフィルタでノイズ除去
        gray = cv2.GaussianBlur(gray, (3, 3), 0)
        
        # 2. 適応的閾値処理（照明ムラに強い）
        bw = cv2.adaptiveThreshold\
            (gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY_INV, 11, 2)
        
        # 3. より大きなカーネルでCLOSE処理
        kernel_close = np.ones((20, 20), np.uint8)
        bw = cv2.morphologyEx(bw, cv2.MORPH_CLOSE, kernel_close)
        
        # 4. OPEN処理でノイズ除去
        kernel_open = np.ones((5, 5), np.uint8)
        bw = cv2.morphologyEx(bw, cv2.MORPH_OPEN, kernel_open)
        
        if debug:
            cv2.imwrite("debug_binary.png", bw)
        
        # 輪郭取得
        contours, _ = cv2.findContours(bw, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        
        h, w = bw.shape
        marks = []
        
        # より柔軟なマージン設定（斜めの画像に対応）
        margin_x = w * 0.15  # 横15%
        margin_y = h * 0.15  # 縦15%
        
        mark_candidates = []
        
        for c in contours:
            area = cv2.contourArea(c)
            
            # 面積フィルタを緩和
            if area < 30 or area > w * h * 0.01:  # 最小30ピクセル、最大画像の1%
                continue
            
            # 輪郭の近似
            epsilon = 0.02 * cv2.arcLength(c, True)
            approx = cv2.approxPolyDP(c, epsilon, True)
            
            x, y, ww, hh = cv2.boundingRect(c)
            cx, cy = x + ww / 2, y + hh / 2
            
            # アスペクト比チェック（正方形に近いものを優先）
            aspect_ratio = ww / hh if hh > 0 else 0
            if not (0.5 <= aspect_ratio <= 2.0):
                continue
            
            # 四隅付近にあるかチェック（より柔軟に）
            is_corner = False
            corner_type = None
            
            # 左上
            if cx < margin_x and cy < margin_y:
                is_corner = True
                corner_type = "top_left"
            # 右上
            elif cx > w - margin_x and cy < margin_y:
                is_corner = True
                corner_type = "top_right"
            # 左下
            elif cx < margin_x and cy > h - margin_y:
                is_corner = True
                corner_type = "bottom_left"
            # 右下
            elif cx > w - margin_x and cy > h - margin_y:
                is_corner = True
                corner_type = "bottom_right"
            
            if is_corner:
                # 品質スコア計算（面積、アスペクト比、輪郭の複雑さを考慮）
                quality_score = area * \
                    (1 / abs(1 - aspect_ratio) if aspect_ratio != 1 else 1) *\
                    (1 / len(approx) if len(approx) > 0 else 0)
                
                mark_candidates.append({
                    'center': (cx, cy),
                    'corner_type': corner_type,
                    'area': area,
                    'quality': quality_score,
                    'aspect_ratio': aspect_ratio
                })
        
        if debug:
            print(f"マーク候補数: {len(mark_candidates)}")
            for i, candidate in enumerate(mark_candidates):
                print(
                    f"候補{i}: {candidate['corner_type']}, "
                    f"中心({candidate['center'][0]:.1f}, {candidate['center'][1]:.1f}), "
                    f"面積:{candidate['area']}, "
                    f"品質:{candidate['quality']:.2f}"
                )
        
        # 各コーナーから最も品質の高い候補を選択
        corner_types = ["top_left", "top_right", "bottom_left", "bottom_right"]
        selected_marks = []
        
        for corner in corner_types:
            candidates = [c for c in mark_candidates if c['corner_type'] == corner]
            if candidates:
                # 品質スコアで並び替えて最良のものを選択
                best = max(candidates, key=lambda x: x['quality'])
                selected_marks.append(best['center'])
        
        if len(selected_marks) == 4:
            # 順序を統一（左上、右上、右下、左下の順）
            return selected_marks
        
        # 4つ見つからない場合は、距離ベースでの補完を試行
        if len(selected_marks) >= 2:
            print(f"Warning: \
                {len(selected_marks)}個のマークのみ検出。距離ベース補完を試行中...")
            # ここで距離や角度から欠けているマークを推定する処理を追加可能
        
        return []
   
    # スキャン画像のマーク検出（デバッグモード有効）
    scanned_bgr = cv2.cvtColor(np.array(scanned_img), cv2.COLOR_RGB2BGR)
    src_pts = find_marks(scanned_bgr, debug=True)
    
    # テンプレート画像のマーク検出
    dst_pts = find_marks(template_img, debug=False)
    
    if len(src_pts) != 4 or len(dst_pts) != 4:
        print(f"Warning: 四隅マークの検出に失敗しました")
        print(f"スキャン画像: {len(src_pts)}個検出")
        print(f"テンプレート: {len(dst_pts)}個検出")
        print("（補正なし）")
        return scanned_img
    
    print("四隅マーク検出成功！射影変換を実行中...")
    
    # numpy配列に変換
    src_pts = np.float32(src_pts)
    dst_pts = np.float32(dst_pts)
    
    # 射影変換行列計算
    M = cv2.getPerspectiveTransform(src_pts, dst_pts)
    
    # 射影変換実行
    h, w = template_img.shape[:2]
    warped = cv2.warpPerspective(np.array(scanned_img), M, (w, h))
    
    return Image.fromarray(warped)


def save_results_to_excel(rows, recipient_digits, name_digits, residence_digits, output_base="output",\
template_xlsx_path=os.path.join(settings.BASE_DIR, "static_files", \
"documents", "output_template.xlsx"), seq_no=1):
# ★変更===template_xlsx_path=os.path.join(settings.BASE_DIR,とする

    # 既存の帳票テンプレートを読み込む
    wb = load_workbook(template_xlsx_path) # ★変更===template_xlsx_pathとした
    ws = wb.active

    # === 固定セル ===
    ws["J5"]  = recipient_digits   # 受給者証番号
    ws["AH5"] = name_digits        # 受給者名（カナ）
    ws["AF7"] = residence_digits   # 住居番号

    # === 行データ ===
    for idx, row in enumerate(rows, start=13):
        ws[f"J{idx}"]  = row.get("Service", "")
        ws[f"R{idx}"]  = row.get("Jutogai", "")
        ws[f"X{idx}"]  = row.get("Night", "")
        ws[f"AB{idx}"]  = row.get("Hospita", "")
        ws[f"AF{idx}"]  = row.get("Gohome", "")
        ws[f"AJ{idx}"] = row.get("Support", "")
        ws[f"AN{idx}"] = row.get("Medic", "")
        ws[f"AR{idx}"] = row.get("Jiritsu1", "")
        ws[f"AV{idx}"] = row.get("Jiritsu2", "")

    # === 新しいファイル名ルール ===
    recipient = (recipient_digits or "").strip()
    filename = f"{recipient}_{seq_no}ページ目.xlsx"

    if os.path.isdir(output_base) or output_base.endswith(('/', '\\')):
        output_path = os.path.join(output_base, filename)
    else:
        output_path = filename

    wb.save(output_path)
    print("Excelに書き込みました ->", output_path)
    return output_path

# ★変更=== 
#    output_path について、output_baseがディレクトリパスの場合はファイル名を結合、
#    そうでなければそのまま使用 ===
    if os.path.isdir(output_base) or output_base.endswith('/') or output_base.endswith('\\'):
        output_path = os.path.join(output_base, filename)
    else:
        # output_baseが既にファイル名を含んでいる場合（従来の使い方との互換性維持）
        output_path = f"{output_base}.xlsx"

    # === 保存 ===
    wb.save(output_path)
    print("Excelに書き込みました ->", output_path)

# ★変更=== ↓新規追加
    # 作成されたファイルのパスを返す
    return output_path


#--- HHMM 正規化（例: "9"→"0900", "10"→"1000", "900"→"0900"）
def normalize_hhmm_digits(digits: str) -> str:
    if not digits:
        return ""
    digits = re.sub(r"\D", "", digits)
    if len(digits) >= 4:
        return digits[:4]
    if len(digits) == 3:
        return digits.zfill(4)  # "900" -> "0900"
    if len(digits) == 2:
        if int(digits) < 24:
            return digits + "00"       # "10" -> "1000"
        else:
            return "0" + digits + "0"  # "90" -> "0900"
    if len(digits) == 1:
        return digits.zfill(2) + "00"  # "9" -> "0900"
    return ""


# -----MODEL_CANDIDATES = ["best_model.h5", "mnist_model.h5"]の読み込み
# -----main内で、model = load_model_if_exists()と宣言
def load_model_if_exists():
    if not _have_tf:
        return None
    for p in MODEL_CANDIDATES:
        if Path(p).exists():
            try:
                print("Loading model:", p)
                model = tf.keras.models.load_model(p)
                print("Loaded model:", p)
                return model
            except Exception as e:
                print("Failed to load model", p, ":", e)
    print("No model found among:", MODEL_CANDIDATES)
    return None


def remove_lines(img, debug=True, debug_name="debug"):
    """
    横幅90%以上・高さ90%以上の線は無条件で白塗り。
    検出条件をシンプルにして確実に線を消す。
    """
    import cv2
    import numpy as np
    from PIL import Image
    import os

    gray = img.convert('L')
    gray_np = np.array(gray)

    # --- Otsuで二値化（線を白背景に黒線にする）
    _, bin_img = cv2.threshold(gray_np, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)

    img_no_lines = gray_np.copy()
    height, width = bin_img.shape

    # --- 横線検出 ---
    horizontal_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (max(3, width // 2), 1))
    detected = cv2.morphologyEx(bin_img, cv2.MORPH_OPEN, horizontal_kernel)
    cnts, _ = cv2.findContours(detected, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    for c in cnts:
        x, y, w_rect, h_rect = cv2.boundingRect(c)
        if w_rect > width * 0.9:  # 横幅90%以上なら絶対白塗り
            cv2.rectangle(img_no_lines, (x, y), (x + w_rect, y + h_rect), 255, -1)

    # --- 縦線検出 ---
    vertical_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, max(3, height // 2)))
    detected = cv2.morphologyEx(bin_img, cv2.MORPH_OPEN, vertical_kernel)
    cnts, _ = cv2.findContours(detected, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    for c in cnts:
        x, y, w_rect, h_rect = cv2.boundingRect(c)
        if h_rect > height * 0.9:  # 高さ90%以上なら絶対白塗り
            cv2.rectangle(img_no_lines, (x, y), (x + w_rect, y + h_rect), 255, -1)

    img_result = Image.fromarray(img_no_lines)

    # --- デバッグ保存 ---
    if debug:
        os.makedirs("remove_line", exist_ok=True)
        Image.fromarray(bin_img).save(f"remove_line/{debug_name}_bin.png")
        img_result.save(f"remove_line/{debug_name}_no_lines.png")

    return img_result


def is_blank_image(img, mode="soft", debug=False, debug_name="debug"):
    """
    空欄判定:
      - remove_lines() で罫線を削除
      - mode="soft": 枠を軽く避ける
      - mode="hard": より厳しく中央寄せ
      - 黒ピクセル割合が閾値以下なら空白
    """
    import os
    from PIL import Image

    # --- 罫線削除 ---
    img_no_lines = remove_lines(img, debug=debug, debug_name=debug_name)

    # --- クロップ＆閾値設定 ---
    w, h = img_no_lines.size
    if mode == "soft":
        threshold = 0.04 #---- 黒が４％以下なら空白
        img_cropped = img_no_lines.crop((int(w*0.015), int(h*0.025),
                                         int(w*0.985), int(h*0.985)))
    else:
        threshold = 0.010
        img_cropped = img_no_lines.crop((int(w*0.1), int(h*0.1),
                                         int(w*0.9), int(h*0.9)))

    # --- 二値化（完全白黒）
    bw = img_cropped.point(lambda x: 0 if x < 150 else 255, '1')

    # --- 黒ピクセル割合計算 ---
    black_pixels = bw.histogram()[0]
    total_pixels = bw.width * bw.height
    black_ratio = black_pixels / total_pixels

    # --- デバッグ出力 ---
    if debug:
        os.makedirs("remove_line", exist_ok=True)
        img_no_lines.save(f"remove_line/{debug_name}_no_lines.png")
        bw.save(f"remove_line/{debug_name}_binary.png")
        print(f"[DEBUG] 黒比率: {black_ratio:.4f} / 閾値: {threshold}")

    return black_ratio <= threshold #--- 画像内の黒ピクセルの割合が閾値以下の場合にTrueを返すということ




# -----モデル入力用に 28x28 -> (1,28,28,1) の numpy 配列にする
def preprocess_for_model(pil_img, invert=True, debug_name="", seq_no=None):
    img = pil_img.convert("L")
    if invert:
        img = ImageOps.invert(img)  # postproc の画像を黒白反転させる

    # パディングして正方形化（余白に黒背景を追加）
    w, h = img.size
    side = max(w, h)
    padded = Image.new("L", (side, side), 0)  # 黒背景
    padded.paste(img, ((side - w) // 2, (side - h) // 2))

    img = padded.resize((28, 28), Image.Resampling.LANCZOS)

    # デバッグ保存
    if debug_name is not None:
        ensure_dir("debug_inputs")
        suffix = f"_p{seq_no}" if seq_no is not None else ""
        img.save(f"debug_inputs/{debug_name}{suffix}_.png")

    arr = np.array(img).astype("float32") / 255.0
    arr = arr.reshape(1, 28, 28, 1)
    return arr


def is_near_large_contour(small_contour, all_contours, keep_flags, distance_threshold=20):
    """小さな輪郭が大きな輪郭の近くにあるかチェック"""
    small_rect = cv2.boundingRect(small_contour)
    small_center = (small_rect[0] + small_rect[2]//2, small_rect[1] + small_rect[3]//2)
    
    for i, large_contour in enumerate(all_contours):
        if keep_flags[i]:  # 既に保持が決まっている輪郭
            large_rect = cv2.boundingRect(large_contour)
            large_center = (large_rect[0] + large_rect[2]//2, large_rect[1] + large_rect[3]//2)
            
            # 距離を計算
            distance = np.sqrt((small_center[0] - large_center[0])**2 + 
                             (small_center[1] - large_center[1])**2)
            
            if distance < distance_threshold:
                return True
    return False


def remove_noise(pil_img, debug_name=""):
    # -----改良版: 小さな '0' を消さないために輪郭判定を賢くする。
    ensure_dir("mid")
    img_gray = pil_img.convert("L")
    if debug_name:
        img_gray.save(Path("mid") / f"{debug_name}_00_original.png")
    arr = np.array(img_gray)
    #--- Otsu の閾値値を取得してマージンを使う
    #--- OpenCV のcv2.threshold：大津の手法 (Otsu's method)
    #---「画像のヒストグラム（二値化前の明るさ分布）を解析して、
    ##--- クラス内分散を最小にするしきい値」を自動的に決める手法
    #--- つまり、明るい背景と暗い文字がある時、それらを最もきれいに分ける閾値
    ##--- （0〜255の数値）を自動で計算してくれる
    otsu_val, _ = cv2.threshold(arr, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    
    thr = max(0, int(otsu_val) - 2)  # margin を少し大きく調整
    _, bw = cv2.threshold(arr, thr, 255, cv2.THRESH_BINARY_INV)
    if debug_name:
        Image.fromarray(bw).save(Path("mid") / f"{debug_name}_01_bw.png")
    
    # 小さなノイズ除去（開閉で軽く掃く）
    bw_open = cv2.morphologyEx(bw, cv2.MORPH_OPEN, np.ones((2,2), np.uint8))
    
    # かすれた部分を補強するためのClosing
    kernel_repair = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (3,3))
    bw_repaired = cv2.morphologyEx(bw_open, cv2.MORPH_CLOSE, kernel_repair)
    
    if debug_name:
        Image.fromarray(bw_open).save(Path("mid") / f"{debug_name}_02_open.png")
        Image.fromarray(bw_repaired).save(Path("mid") / f"{debug_name}_02_repaired.png")  # 修復版も保存
    
    # 輪郭検出は修復版から（重複削除）
    contours, hierarchy = cv2.findContours(bw_repaired.copy(), cv2.RETR_CCOMP, cv2.CHAIN_APPROX_SIMPLE)
    
    h_img, w_img = bw_repaired.shape  # bw_openではなくbw_repairedのサイズを使用
    clean_mask = np.zeros_like(bw_repaired)
    
    # 判定用パラメータ（「３」の上部も保持するように調整）
    AREA_LARGE = 50       # 少し下げる（60→50）
    AREA_SMALL = 8        # 少し下げる（10→8）
    H_RATIO = 0.20        # 少し下げる（0.25→0.20）
    H_SMALL_RATIO = 0.06  # 少し下げる（0.10→0.06）
    MIN_W_RATIO = 0.025   # 少し下げる（0.03→0.025）
    
    if hierarchy is None:
        hierarchy = np.array([]).reshape(1,0,4)
    
    # まず「保持候補」を集める
    keep_flags = [False] * len(contours)
    for i, c in enumerate(contours):
        area = cv2.contourArea(c)
        x, y, w, h = cv2.boundingRect(c)
        ar = w / float(h + 1e-9)
        keep = False
        
        # 大きくて高さ比が取れるなら確定
        if area >= AREA_LARGE and h >= h_img * H_RATIO:
            keep = True
        else:
            # 小さめでも縦横比・高さで候補
            if area >= AREA_SMALL and (h >= h_img * H_SMALL_RATIO or w >= w_img * MIN_W_RATIO):
                if 0.15 <= ar <= 6.0:
                    keep = True
            
            # 横線用の追加条件
            if not keep and ar > 1.5 and w >= w_img * 0.10:  # 少し緩和（0.12→0.10）
                keep = True
        
        # 親子関係がある場合（穴や内部構造）は保持
        if not keep:
            child_idx = hierarchy[0][i][2] if hierarchy.size else -1
            parent_idx = hierarchy[0][i][3] if hierarchy.size else -1
            if child_idx != -1 or parent_idx != -1:
                keep = True
        
        keep_flags[i] = keep
    
    # 描画：保持フラグが True な輪郭を塗る
    for i, c in enumerate(contours):
        if keep_flags[i]:
            cv2.drawContours(clean_mask, [c], -1, 255, thickness=-1)
    
    if debug_name:
        Image.fromarray(clean_mask).save(Path("mid") / f"{debug_name}_03_mask.png")
    
    # 子輪郭（穴）を黒で描画して穴を作る（parent/child 情報を利用）
    for i, c in enumerate(contours):
        parent_idx = hierarchy[0][i][3] if hierarchy.size else -1
        if parent_idx != -1:
            # 親が存在するなら穴として残す（親が描かれていれば効果あり）
            cv2.drawContours(clean_mask, [c], -1, 0, thickness=-1)
    
    if debug_name:
        Image.fromarray(clean_mask).save(Path("mid") / f"{debug_name}_03_mask_after_holes.png")
    
    # 最終反転（モデルは白背景黒文字想定）
    final_mask = cv2.bitwise_not(clean_mask)
    final_img = Image.fromarray(final_mask)
    if debug_name:
        final_img.save(Path("mid") / f"{debug_name}_04_final.png")
    
    return final_img



def split_into_digits2(img, debug_name=""):
    # -----OpenCVで桁分割（debug_name 指定で mid/ に途中画像を保存）
    if not _have_cv2:
        return [img]

    # コロン除去＋ノイズ処理
    img_clean = remove_noise(img, debug_name)

    # === 枠を避けるために少し内側をトリミング ===
    trim_x = int(img_clean.width * 0.04)   # 左右4%カット（調整可）
    trim_y = int(img_clean.height * 0.06)  # 上下6%カット（調整可）
    img_proc = img_clean.crop((
        trim_x,
        trim_y,
        img_clean.width - trim_x,
        img_clean.height - trim_y
    ))
    if debug_name:
        ensure_dir("mid")
        img_proc.save(Path("mid") / f"{debug_name}_trimmed.png")

    return img_proc



def predict_digits_with_model(model, img, debug_name=""):
    if model is None:
        return ""

    # 空欄判定（画像全体）
    #if is_blank_image(img):
    #    return ""
    ### ここの#をとった場合、食事加算で一部の「1」を読み込まない→エクセルには何も出力されない
    ### ここに#を付けると、読み込まなかった「1」が、エクセルで「0」が出力された
    ### 何も出力されないよりはマシという判断



    # 桁分割（debug_name を渡す）
    imgs = split_into_digits2(img, debug_name) # 1830のような時刻画像を1,8,3,0,のように切り出す。

    preds = []

    # debug 保存 (桁ごとの中身を確認したい時)
    if debug_name:
        ensure_dir("mid")
        imgs.save(Path("mid") / f"{debug_name}_postproc.png")

    arr = preprocess_for_model(imgs, invert=True, debug_name=debug_name)
    out = model.predict(arr, verbose=0)
    d = str(int(np.argmax(out, axis=1)[0]))
    if d.isdigit():
        preds.append(d)

    result = "".join(preds)
    result = result.replace(":", "").replace("・", "").replace(".", "")

    return result


# サービス提供状況の数字を対応する文字列に変換する関数
def convert_service_value(raw_value):
    """
    サービス提供状況の数字を対応する文字列に変換
    1-8以外の値は空文字列を返す
    """
    # 空文字列や空白の場合
    if not raw_value or not raw_value.strip():
        return ""
    
    # 数字のみを抽出（念のため）
    digits = re.sub(r"\D", "", raw_value.strip())
    
    # 変換辞書
    service_mapping = {
        "1": "1. 入院（本体報酬）",
        "2": "2. 外泊（本体報酬）",
        "3": "3. 入院",
        "4": "4. 外泊",
        "5": "5. 入院→ 外泊",
        "6": "6. 外泊→入院",
        "7": "7. 入院→共同生活住居に戻る→外泊",
        "8": "8. 外泊→共同生活住居に戻る→入院"
    }
    
    # 該当する値があれば変換、なければ空文字列
    return service_mapping.get(digits, "")


#--- 開始・終了時刻"以外"を OCR → fallback モデル の2段階で推定する
def ocr_etc_with_fallback(pil_img, model=None, debug_name=""):
    """時刻以外の数字欄OCR → モデルのみ"""
    if model is not None:
        val = predict_digits_with_model(model, pil_img, debug_name=debug_name)
        return val
    return ""


# ★変更=== viewsからpredict_from_pdf_B.pyを呼ぶための処理　mainをprocess_pdf_to_excelに変更
from datetime import datetime

# ------------------ 実行本体 ------------------
def process_pdf_to_excel_with_progress(pdf_path, output_dir="downloads", progress_callback=None):
    """
    進捗コールバック付きのPDF変換処理
    progress_callback: function(current_page, total_pages, message)
    """
    ensure_dir(output_dir)
    model = load_model_if_exists()
    output_files = []
    doc = fitz.open(pdf_path)
    total_pages = doc.page_count

    # 進捗コールバック実行用ヘルパー
    def update_progress(current, message):
        if progress_callback:
            # 整数に変換して渡す
            current_int = int(round(current)) if isinstance(current, float) else current
            total_int = int(total_pages)
            progress_callback(current_int, total_int, message)

    try:
        update_progress(0, "PDFファイルを読み込み中...")
        
        # === ページごとに処理 ===
        for seq_no, page_index in enumerate(range(doc.page_count), start=1):
            
            page_img = pdf_page_to_pil(pdf_path, page_index, DPI)
            page_img = align_using_corner_marks(page_img, DPI, 
                            os.path.join(settings.BASE_DIR, "static_files", "documents", "template.pdf"))

            update_progress(seq_no, f"{seq_no}ページ目のOCR処理中...")


            # 1) 受給者証番号
            recipient_box = crop_norm(page_img, BOX_RECIPIENT)
            #recipient_box.save(Path(DEBUG_DIR) / "recipient_box.png")
            recipient_digits = predict_recipient_number(recipient_box, model, debug_name=f"recipient_{seq_no}")

            # 2) 受給者名（カナ）
            name_box = crop_norm(page_img, BOX_NAME)
            #name_box.save(Path(DEBUG_DIR) / "name_box.png")
            name_digits = predict_name(name_box, model, debug_name=f"name_{seq_no}")

            # 3) 住居番号
            residence_box = crop_norm(page_img, BOX_RESIDENCE)
            #residence_box.save(Path(DEBUG_DIR) / "residence_box.png")
            residence_digits = predict_residence_number(residence_box, model, debug_name=f"residence_{seq_no}")

            # ) 各列を分割
            service_box  = crop_norm(page_img, BOX_SERVICE)
            jutogai_box   = crop_norm(page_img, BOX_JUTOGAI)
            nights_box     = crop_norm(page_img, BOX_NIGHT)
            hospita_box   = crop_norm(page_img, BOX_HOSPITA)
            gohome_box = crop_norm(page_img, BOX_GOhome)
            support_box     = crop_norm(page_img, BOX_SUPPORT)
            medic_box    = crop_norm(page_img, BOX_MEDIC)
            jiritsu1_box    = crop_norm(page_img, BOX_JIRITSU1)
            jiritsu2_box     = crop_norm(page_img, BOX_JIRITSU2)

        #ensure_dir(os.path.join(DEBUG_DIR, "services"))
        #ensure_dir(os.path.join(DEBUG_DIR, "jutogai"))
        #ensure_dir(os.path.join(DEBUG_DIR, "nights"))
        #ensure_dir(os.path.join(DEBUG_DIR, "hospita"))
        #ensure_dir(os.path.join(DEBUG_DIR, "gohome"))
        #ensure_dir(os.path.join(DEBUG_DIR, "support"))
        #ensure_dir(os.path.join(DEBUG_DIR, "medics"))
        #ensure_dir(os.path.join(DEBUG_DIR, "jiritsu1"))
        #ensure_dir(os.path.join(DEBUG_DIR, "jiritsu2"))

    # === 行分割（まず罫線ベース、自動失敗時は均等割にフォールバック） ===
            def split_rows(box_img, n_rows, label):
                rows = split_rows_by_lines(box_img, debug_name=f"{label}_{seq_no}")
                if len(rows) != n_rows:
                    print(f"Warning: {label} 検出行数={len(rows)}。均等分割にフォールバック。")
                    rows = split_vertical(box_img, n_rows)
                return rows

    # === 実際に分割を実行 ===　★↑のdef split_rows(box_img, n_rows, label)に当てはめている
            service_rows  = split_rows(service_box, NUM_ROWS, "services")
            jutogai_rows   = split_rows(jutogai_box, NUM_ROWS, "jutogai")
            nights_rows     = split_rows(nights_box, NUM_ROWS, "nights")
            hospita_rows   = split_rows(hospita_box, NUM_ROWS, "hospita")
            gohome_rows = split_rows(gohome_box, NUM_ROWS, "gohome")
            support_rows     = split_rows(support_box, NUM_ROWS, "support")
            medic_rows    = split_rows(medic_box, NUM_ROWS, "medics")
            jiritsu1_rows    = split_rows(jiritsu1_box, NUM_ROWS, "jiritsu1")
            jiritsu2_rows     = split_rows(jiritsu2_box, NUM_ROWS, "jiritsu2")

        # === 行ごとの処理 ===
            rows = []
            for i in range(NUM_ROWS):
                day = i + 1
                v_img = service_rows[i]
                j_img = jutogai_rows[i]
                n_img = nights_rows[i]
                h_img = hospita_rows[i]
                g_img = gohome_rows[i]
                s_img = support_rows[i]
                c_img = medic_rows[i]
                j1_img = jiritsu1_rows[i]
                j2_img = jiritsu2_rows[i]

            #v_img.save(Path(DEBUG_DIR) / "services" / f"service_{day:02d}.png")
            #j_img.save(Path(DEBUG_DIR) / "jutogai" / f"jutogai_{day:02d}.png")
            #n_img.save(Path(DEBUG_DIR) / "nights" / f"night_{day:02d}.png")
            #h_img.save(Path(DEBUG_DIR) / "hospita" / f"hospita_{day:02d}.png")
            #g_img.save(Path(DEBUG_DIR) / "gohome" / f"gohome_{day:02d}.png")
            #s_img.save(Path(DEBUG_DIR) / "support" / f"support_{day:02d}.png")
            #c_img.save(Path(DEBUG_DIR) / "medics" / f"medic_{day:02d}.png")
            #j1_img.save(Path(DEBUG_DIR) / "jiritsu1" / f"jiritsu1_{day:02d}.png")
            #j2_img.save(Path(DEBUG_DIR) / "jiritsu2" / f"jiritsu2_{day:02d}.png")

    # 予測　ocr_time_with_fallback・ocr_etc_with_fallback関数に投げる
        #---サービス提供の状況
                if is_blank_image(v_img, "hard", debug_name=f"service_{seq_no}_{day:02d}"):
                    service_val = ""
                else:
                    raw_service = ocr_etc_with_fallback(v_img, model, debug_name=f"service_{seq_no}_{day:02d}")
                    service_val = convert_service_value(raw_service)
        #---住登外
                jutogai_val = "" if is_blank_image\
                    (j_img, "hard", debug_name=f"jutogai_{seq_no}_{day:02d}") else "1"

        #---夜間支援
                night_val = "" if is_blank_image(n_img, "hard", \
                    debug_name=f"night_{seq_no}_{day:02d}") else\
                    ocr_etc_with_fallback(n_img, model, debug_name=f"night_{seq_no}_{day:02d}")
        #---入院
                hospita_val = "" if is_blank_image(h_img, "hard", debug_name=f"hospita_{seq_no}_{day:02d}") else "1"
        #---帰宅
                gohome_val = "" if is_blank_image(g_img, "hard", debug_name=f"gohome_{seq_no}_{day:02d}") else "1"
        #---日中支援加算 ハードモードで空白判定
                support_val = "" if is_blank_image(s_img, "hard", debug_name=f"support_{seq_no}_{day:02d}") else "1"
        #---医療連携体制加算
                medic_val = "" if is_blank_image(c_img, "hard", \
                    debug_name=f"medic_{seq_no}_{day:02d}") else\
                    ocr_etc_with_fallback(c_img, model, debug_name=f"medic_{seq_no}_{day:02d}")
        #---自立支援加算（Ⅰ）
                jiritsu1_val   = "" if is_blank_image(j1_img, "hard", debug_name=f"jiritsu1_{seq_no}_{day:02d}") else "1"

        #---自立支援加算（Ⅱ）
                jiritsu2_val    = "" if is_blank_image(j2_img, "hard", \
                    debug_name=f"jiritsu2_{seq_no}_{day:02d}") else "1"

                rows.append({
                    "Day": day,
                    "Service": service_val,
                    "Jutogai": jutogai_val,
                    "Night": night_val,
                    "Hospita": hospita_val,
                    "Gohome": gohome_val,
                    "Support": support_val,
                    "Medic": medic_val,
                    "Jiritsu1": jiritsu1_val,
                    "Jiritsu2": jiritsu2_val,
                    "RecipientNumber": recipient_digits,
                    "Name": name_digits,
                    "ResidenceNumber": residence_digits,
                })

            # Excelファイル作成
            created_file = save_results_to_excel(
                rows=rows,
                recipient_digits=recipient_digits,
                name_digits=name_digits,
                residence_digits=residence_digits,
                output_base=output_dir,
                template_xlsx_path=os.path.join(settings.BASE_DIR, "static_files", "documents", "output_template.xlsx"),
                seq_no=seq_no
            )
            output_files.append(created_file)
            
            update_progress(seq_no, f"{seq_no}ページ目の処理完了")

        update_progress(total_pages, "全ての処理が完了しました")

    except Exception as e:
        update_progress(0, f"エラーが発生しました: {str(e)}")
        raise e
    finally:
        doc.close()

    return output_files


# ★修正=== 開発用: ターミナルでpython predict_from_pdf_B.pyで動かす時
if __name__ == "__main__":
    files = process_pdf_to_excel(PDF_PATH)
    print("生成したファイル:", files)
